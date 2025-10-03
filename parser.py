import json
import sys
import requests
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from urllib3.exceptions import InsecureRequestWarning
import os.path
import time
from bs4 import BeautifulSoup
import re
from datetime import datetime
import os
import html

# Disable SSL warnings
requests.packages.urllib3.disable_warnings(category=InsecureRequestWarning)


def get_language_from_extension(file_path):
    """Determine programming language based on file extension"""
    if not file_path:
        return "Unknown"

    extension = os.path.splitext(file_path)[1].lower()

    language_map = {
        ".java": "Java",
        ".py": "Python",
        ".js": "JavaScript",
        ".ts": "TypeScript",
        ".jsx": "React JSX",
        ".tsx": "React TSX",
        ".cpp": "C++",
        ".c": "C",
        ".cs": "C#",
        ".php": "PHP",
        ".rb": "Ruby",
        ".go": "Go",
        ".rs": "Rust",
        ".swift": "Swift",
        ".kt": "Kotlin",
        ".scala": "Scala",
        ".html": "HTML",
        ".css": "CSS",
        ".xml": "XML",
        ".json": "JSON",
        ".yaml": "YAML",
        ".yml": "YAML",
        ".sql": "SQL",
        ".sh": "Shell",
        ".bat": "Batch",
        ".ps1": "PowerShell",
        ".md": "Markdown",
        ".dockerfile": "Docker",
        ".tf": "Terraform",
        ".jsn": "JSON",
    }

    return language_map.get(extension, "Unknown")


def extract_filename(component):
    """Extract filename from component field (after colon)"""
    if not component:
        return ""

    if ":" in component:
        return component.split(":", 1)[1]
    return component


def generate_single_html_report(
    project_name,
    project_version,
    all_issues,
    output_filename="sonarqube_issues_report.html",
):
    """
    Generate a single HTML report with all issues and source code snippets
    with syntax highlighting using Highlight.js
    """
    try:
        # Generate comprehensive statistics
        def generate_statistics():
            severity_counts = {}
            status_counts = {}
            type_counts = {
                "VULNERABILITY": len(
                    [
                        i
                        for i in all_issues
                        if i.get("type", "").upper() == "VULNERABILITY"
                    ]
                ),
                "BUG": len(
                    [i for i in all_issues if i.get("type", "").upper() == "BUG"]
                ),
                "CODE_SMELL": len(
                    [i for i in all_issues if i.get("type", "").upper() == "CODE_SMELL"]
                ),
            }

            for issue in all_issues:
                severity = issue.get("severity", "UNKNOWN")
                status = issue.get("status", "UNKNOWN")

                severity_counts[severity] = severity_counts.get(severity, 0) + 1
                status_counts[status] = status_counts.get(status, 0) + 1

            return severity_counts, status_counts, type_counts

        severity_counts, status_counts, type_counts = generate_statistics()
        total_issues = len(all_issues)

        # Categorize issues
        vulnerabilities = [
            issue
            for issue in all_issues
            if issue.get("type", "").upper() == "VULNERABILITY"
        ]
        bugs = [issue for issue in all_issues if issue.get("type", "").upper() == "BUG"]
        code_smells = [
            issue
            for issue in all_issues
            if issue.get("type", "").upper() == "CODE_SMELL"
        ]

        # HTML template with CSS styling and Highlight.js syntax highlighting
        html_template = f"""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{project_name} - SonarQube Security Report</title>
    
    <!-- Highlight.js for syntax highlighting -->
    <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/styles/github-dark.min.css">
    <script src="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/highlight.min.js"></script>
    
    <!-- Load additional languages -->
    <script src="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/languages/java.min.js"></script>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/languages/python.min.js"></script>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/languages/javascript.min.js"></script>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/languages/typescript.min.js"></script>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/languages/cpp.min.js"></script>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/languages/csharp.min.js"></script>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/languages/php.min.js"></script>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/languages/ruby.min.js"></script>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/languages/go.min.js"></script>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/languages/rust.min.js"></script>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/languages/swift.min.js"></script>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/languages/kotlin.min.js"></script>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/languages/scala.min.js"></script>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/languages/css.min.js"></script>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/languages/sql.min.js"></script>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/languages/bash.min.js"></script>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/languages/yaml.min.js"></script>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/languages/json.min.js"></script>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/languages/markdown.min.js"></script>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/languages/xml.min.js"></script>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/languages/dockerfile.min.js"></script>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/languages/plaintext.min.js"></script>
    
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            line-height: 1.6;
            color: #333;
            background-color: #f8f9fa;
        }}
        
        .container {{
            max-width: 1200px;
            margin: 0 auto;
            padding: 20px;
        }}
        
        /* Header Styles */
        .header {{
            background: linear-gradient(135deg, #2C3E50, #4CA1AF);
            color: white;
            padding: 40px 0;
            text-align: center;
            margin-bottom: 30px;
            border-radius: 10px;
            box-shadow: 0 4px 6px rgba(0,0,0,0.1);
        }}
        
        .header h1 {{
            font-size: 2.5em;
            margin-bottom: 10px;
            font-weight: 300;
        }}
        
        .header .subtitle {{
            font-size: 1.2em;
            opacity: 0.9;
            margin-bottom: 15px;
            font-size: 2em;
        }}
        
        .header .project-version {{
            margin-bottom: 15px;
            opacity: 0.9;
        }}
        
        .header .date {{
            font-size: 0.9em;
            opacity: 0.8;
        }}
        
        /* Summary Cards */
        .summary-cards {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
            gap: 20px;
            margin-bottom: 30px;
        }}
        
        .card {{
            background: white;
            padding: 25px;
            border-radius: 10px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
            text-align: center;
            transition: transform 0.3s ease;
        }}
        
        .card:hover {{
            transform: translateY(-5px);
        }}
        
        .card h3 {{
            color: #2C3E50;
            margin-bottom: 10px;
            font-size: 1.1em;
        }}
        
        .card .count {{
            font-size: 2.5em;
            font-weight: bold;
            margin-bottom: 5px;
        }}
        
        .card.total .count {{ color: #2C3E50; }}
        .card.vulnerabilities .count {{ color: #E74C3C; }}
        .card.bugs .count {{ color: #F39C12; }}
        .card.code-smells .count {{ color: #3498DB; }}
        
        /* Statistics Tables */
        .statistics {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(300px, 1fr));
            gap: 20px;
            margin-bottom: 30px;
        }}
        
        .stat-table {{
            background: white;
            border-radius: 10px;
            overflow: hidden;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        
        .stat-table h3 {{
            background: #34495E;
            color: white;
            padding: 15px 20px;
            margin: 0;
            font-size: 1.1em;
        }}
        
        .stat-table table {{
            width: 100%;
            border-collapse: collapse;
        }}
        
        .stat-table th,
        .stat-table td {{
            padding: 12px 15px;
            text-align: left;
            border-bottom: 1px solid #ECF0F1;
        }}
        
        .stat-table th {{
            background: #ECF0F1;
            font-weight: 600;
            color: #2C3E50;
        }}
        
        .stat-table tr:hover {{
            background: #F8F9FA;
        }}
        
        /* Issues Section */
        .category-section {{
            margin-bottom: 40px;
        }}
        
        .category-header {{
            background: linear-gradient(135deg, #34495E, #2C3E50);
            color: white;
            padding: 20px;
            border-radius: 10px;
            margin-bottom: 10px;
        }}
        
        .issue-card {{
            background: white;
            margin-bottom: 20px;
            border-radius:10px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
            overflow: hidden;
        }}
        
        .issue-header {{
            background: #a1caf9;
            color: #2C3E50;
            padding: 15px 20px;
            font-weight: bold;
            padding-left: 0;
            border-bottom: 0;
        }}
        
        .issue-number {{
            background: #ffffff00;
            color: #f4f8fb;
            padding: 10px 15px;
            font-weight: bold;
            border-radius: 11px;
            margin-left: 15px;
            font-size: 1.5rem;
        }}
        
        .issue-source-file {{
            font-size: 1.1rem;
            padding: 5px 5px;
            background: #f4f8fb;
            color: #2C3E50;
            padding: 10px 15px;
            font-weight: bold;
            border-radius: 17px;
            margin-left: 15px;
            margin-top: 5px;
            max-width: 800px;
            word-break: break-all;
        }}
        
        .issue-details {{
            padding: 20px;
        }}
        
        /* Enhanced Error Message Block */
        .error-message-block {{
            border-radius: 8px;
            padding: 20px;
            margin: 15px 0;
            background: #c3080808;
            border: 1px solid #ffc5c5;
            border-radius: 5px;
        }}
        
        .error-message-header {{
            display: flex;
            align-items: center;
            margin-bottom: 0;
        }}
        
        .error-message-title {{
            font-weight: bold;
            font-size: 1.1em;
        }}
        
        .error-message-content {{
            color: #3a3a3a;
            font-size: 1.1em;
            line-height: 1.5;
            padding: 15px 0;
            border-radius: 5px;
        }}
        
        .detail-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 15px;
            padding: 15px;
            border: 1px solid #d5d5d5;
            border-radius: 5px;
        }}
        
        .detail-item {{
            margin-bottom: 8px;
            border-right: 1px solid #d5d5d5;
            padding: 0 5px;
        }}
        .detail-item:last-child {{
            margin-bottom: 8px;
            border-right: none;
        }}
        
        .detail-label {{
            font-weight: bold;
            color: #2C3E50;
            margin-bottom: 3px;
        }}
        
        .detail-value {{
            color: #555;
            word-wrap: break-word;
        }}
        
        /* Enhanced Source Code with Highlight.js Syntax Highlighting */
        .source-code {{
            background: #2d2d2d;
            color: #f8f8f2;
            border-radius: 8px;
            overflow: hidden;
            margin: 20px 0;
            font-family: 'Consolas', 'Monaco', 'Courier New', monospace;
            box-shadow: 0 4px 6px rgba(0,0,0,0.3);
        }}
        
        .code-header {{
            background: #1a1a1a;
            color: #f8f8f2;
            padding: 12px 20px;
            font-weight: bold;
            border-bottom: 1px solid #444;
            display: flex;
            justify-content: space-between;
            align-items: center;
        }}
        
        .code-language {{
            background: #E74C3C;
            color: white;
            padding: 4px 8px;
            border-radius: 4px;
            font-size: 0.8em;
        }}
        
        .code-content {{
            padding: 0;
            overflow-x: auto;
        }}
        
        .code-table {{
            width: 100%;
            border-collapse: collapse;
        }}
        
        .code-table td {{
            padding: 6px 12px;
            vertical-align: top;
            border: none;
            font-family: 'Consolas', 'Monaco', 'Courier New', monospace;
            font-size: 13px;
            line-height: 1.4;
        }}
        
        .line-number {{
            background: #1a1a1a;
            color: #6c6c6c;
            text-align: right;
            width: 40px;
            border-right: 1px solid #444;
            user-select: none;
            padding-right: 15px;
        }}
        
        .line-content {{
            white-space: normal;
            word-break: keep-all;
            overflow-wrap: normal;
            padding-left: 15px;
        }}
        
        .highlighted-line {{
            background: #3a3a3a !important;
            border-left: 4px solid #569CD6;
        }}
        
        .highlighted-line .line-number {{
            background: #3a3a3a !important;
            color: #569CD6;
            font-weight: bold;
        }}
        
        /* Highlight.js overrides for better integration */
        .hljs {{
            background: transparent !important;
            padding: 0 !important;
        }}
        
        .source-code pre {{
            margin: 0 !important;
            background: transparent !important;
        }}
        
        .source-code code {{
            background: transparent !important;
            font-family: 'Consolas', 'Monaco', 'Courier New', monospace !important;
            font-size: 13px !important;
            line-height: 1.4 !important;
        }}
        
        /* Comments */
        .comment {{
            border: 1px solid #2ecc71;
            border-radius: 5px;
            padding: 15px;
            margin: 10px 0;
            background: #fafdfbad;
        }}
        
        .comment-header {{
            font-weight: bold;
            color: #2C3E50;
            margin-bottom: 8px;
            display: flex;
            justify-content: space-between;
        }}
        
        .comment-content {{
            color: #555;
            line-height: 1.5;
        }}
        
        /* Severity badges */
        .severity-badge {{
            display: inline-block;
            padding: 3px 8px;
            border-radius: 12px;
            font-size: 0.8em;
            font-weight: bold;
            margin-left: 10px;
        }}
        
        .severity-BLOCKER {{ background: #E74C3C; color: white; }}
        .severity-CRITICAL {{ background: #E67E22; color: white; }}
        .severity-MAJOR {{ background: #F39C12; color: white; }}
        .severity-MINOR {{ background: #3498DB; color: white; }}
        .severity-INFO {{ background: #2ECC71; color: white; }}
        
        /* Status badges */
        .status-badge {{
            display: inline-block;
            padding: 3px 8px;
            border-radius: 12px;
            font-size: 0.8em;
            font-weight: bold;
            margin-left: 10px;
        }}
        
        .status-OPEN {{ background: #E74C3C; color: white; }}
        .status-CONFIRMED {{ background: #3498DB; color: white; }}
        .status-REOPENED {{ background: #9B59B6; color: white; }}
        .status-RESOLVED {{ background: #2ECC71; color: white; }}
        .status-CLOSED {{ background: #95A5A6; color: white; }}
        
        /* Print styles */
        @media print {{
            body {{ background: white; }}
            .header {{ background: #2C3E50 !important; }}
            .card {{ break-inside: avoid; }}
            .issue-card {{ break-inside: avoid; }}
        }}
    </style>
</head>
<body>
    <div style="position: fixed; top: 20px; right: 20px; z-index: 1000;">
        <button onclick="switchLanguage()" style="padding: 10px 15px; background: #34495E; color: white; border: none; border-radius: 5px; cursor: pointer; font-size: 14px;">
            Русский / English
        </button>
    </div>
    <div class="container">
        <!-- Header Section -->
        <div class="header">
            <h1>{project_name}</h1>
            <div class="project-version">version {project_version}</div>
            <div class="subtitle" data-i18n-key="report-title">SonarQube Security Report</div>
            <div class="date"><span data-i18n-key="report-date-title">Generated on:</span> {datetime.now().strftime('%Y-%m-%d %H:%M')}</div>
        </div>
        
        <!-- Summary Cards -->
        <div class="summary-cards">
            <div class="card total">
                <h3 data-i18n-key="total-issues">Total Issues</h3>
                <div class="count">{total_issues}</div>
            </div>
            <div class="card vulnerabilities">
                <h3 data-i18n-key="vulnerabilities">Vulnerabilities</h3>
                <div class="count">{type_counts['VULNERABILITY']}</div>
            </div>
            <div class="card bugs">
                <h3 data-i18n-key="bugs">Bugs</h3>
                <div class="count">{type_counts['BUG']}</div>
            </div>
            <div class="card code-smells">
                <h3 data-i18n-key="code-smells">Code Smells</h3>
                <div class="count">{type_counts['CODE_SMELL']}</div>
            </div>
        </div>
        
        <!-- Statistics Section -->
        <div class="statistics">
            <!-- Severity Distribution -->
            <div class="stat-table">
                <h3 data-i18n-key="severity-distribution">Severity Distribution</h3>
                <table>
                    <thead>
                        <tr>
                            <th data-i18n-key="severity">Severity</th>
                            <th data-i18n-key="count">Count</th>
                            <th data-i18n-key="percentage">Percentage</th>
                        </tr>
                    </thead>
                    <tbody>
        """

        # Add severity rows
        severity_order = ["BLOCKER", "CRITICAL", "MAJOR", "MINOR", "INFO"]
        for severity in severity_order:
            if severity in severity_counts:
                count = severity_counts[severity]
                percentage = (count / total_issues) * 100 if total_issues else 0
                html_template += f"""
                        <tr>
                            <td data-i18n-key="severity-{severity}">{severity}</td>
                            <td>{count}</td>
                            <td>{percentage:.1f}%</td>
                        </tr>
                """

        html_template += """
                    </tbody>
                </table>
            </div>
            
            <!-- Status Distribution -->
            <div class="stat-table">
                <h3 data-i18n-key="status-distribution">Status Distribution</h3>
                <table>
                    <thead>
                        <tr>
                            <th data-i18n-key="status">Status</th>
                            <th data-i18n-key="count">Count</th>
                            <th data-i18n-key="percentage">Percentage</th>
                        </tr>
                    </thead>
                    <tbody>
        """

        # Add status rows
        for status, count in sorted(status_counts.items()):
            percentage = (count / total_issues) * 100 if total_issues else 0
            html_template += f"""
                        <tr>
                            <td data-i18n-key="status-{status}">{status}</td>
                            <td>{count}</td>
                            <td>{percentage:.1f}%</td>
                        </tr>
            """

        html_template += """
                    </tbody>
                </table>
            </div>
        </div>
        
        <!-- Detailed Issues Section -->
        <div class="issues-section">
        """

        # Process each category
        categorized_issues = {
            "VULNERABILITY": vulnerabilities,
            "BUG": bugs,
            "CODE_SMELL": code_smells,
        }

        category_names = {
            "VULNERABILITY": "VULNERABILITIES",
            "BUG": "BUGS",
            "CODE_SMELL": "CODE SMELLS",
        }

        # Language to Highlight.js mapping
        hljs_language_map = {
            "Java": "java",
            "Python": "python",
            "JavaScript": "javascript",
            "TypeScript": "typescript",
            "React JSX": "javascript",  # Highlight.js uses jsx for JSX
            "React TSX": "typescript",  # Highlight.js uses tsx for TSX
            "C++": "cpp",
            "C": "c",
            "C#": "csharp",
            "PHP": "php",
            "Ruby": "ruby",
            "Go": "go",
            "Rust": "rust",
            "Swift": "swift",
            "Kotlin": "kotlin",
            "Scala": "scala",
            "HTML": "html",
            "CSS": "css",
            "XML": "xml",
            "JSON": "json",
            "YAML": "yaml",
            "SQL": "sql",
            "Shell": "bash",
            "Batch": "batch",
            "PowerShell": "powershell",
            "Markdown": "markdown",
            "Docker": "dockerfile",
            "Terraform": "hcl",  # Terraform uses HCL syntax
            "Unknown": "plaintext",
        }

        for category, issues in categorized_issues.items():
            if not issues:
                continue

            html_template += f"""
            <div class="category-section">
                <h2 class="category-header"><span data-i18n-key="category-header-{category}">{category_names[category]}</span> ({len(issues)})</h2>
            """

            # Process each issue in this category
            for i, issue in enumerate(issues, 1):
                component = issue.get("component", "")
                file_path = extract_filename(component)
                line_info = issue.get("textRange", {})
                start_line = line_info.get("startLine", "")
                status = issue.get("status", "")
                severity = issue.get("severity", "")
                message = html.escape(issue.get("message", ""))

                language = get_language_from_extension(file_path)
                hljs_lang = hljs_language_map.get(language, "plaintext")

                html_template += f"""
                <div class="issue-card">
                    <div class="issue-header">
                        <span class="issue-number">{i}</span> <span class="issue-source-file">{file_path} : <span>{start_line}</span></span>
                        <span class="severity-badge severity-{severity}" data-i18n-key="severity-{severity}">{severity}</span>
                        <span class="status-badge status-{status}" data-i18n-key="status-{status}">{status}</span>
                    </div>
                    
                    <div class="issue-details">
                        <!-- Enhanced Error Message Block -->
                        
                        <div class="detail-grid">
                            <div class="detail-item">
                                <div class="detail-label" data-i18n-key="key">Key</div>
                                <div class="detail-value">{issue.get('key', 'N/A')}</div>
                            </div>
                            <div class="detail-item">
                                <div class="detail-label" data-i18n-key="type">Type</div>
                                <div class="detail-value">{issue.get('type', 'N/A')}</div>
                            </div>
                            <div class="detail-item">
                                <div class="detail-label" data-i18n-key="author">Author</div>
                                <div class="detail-value">{issue.get('author', 'N/A')}</div>
                            </div>
                            <div class="detail-item">
                                <div class="detail-label" data-i18n-key="rule">Rule</div>
                                <div class="detail-value">{issue.get('rule', 'N/A')}</div>
                            </div>
                            <div class="detail-item">
                                <div class="detail-label" data-i18n-key="language">Language</div>
                                <div class="detail-value">{language}</div>
                            </div>
                        </div>
                        
                        <div class="error-message-block">
                            <div class="error-message-header">
                                <span class="error-message-title" data-i18n-key="issue-description">Issue Description</span>
                            </div>
                            <div class="error-message-content">
                                {message}
                            </div>
                        </div>
                """

                # Source code section with Highlight.js syntax highlighting
                sources = issue.get("sources", [])
                if sources:
                    html_template += f"""
                        <div class="source-code">
                            <div class="code-header">
                                <span data-i18n-key="source-code">Source Code</span>
                                <span class="code-language">{language}</span>
                            </div>
                            <div class="code-content">
                                <table class="code-table">
                    """

                    highlight_line = str(start_line)

                    for source in sources[:15]:  # Limit to first 15 lines
                        line_num = str(source.get("line", ""))
                        code = source.get("code", "")

                        line_class = (
                            "highlighted-line" if line_num == highlight_line else ""
                        )

                        html_template += f"""
                                <tr class="{line_class}">
                                    <td class="line-number">{line_num}</td>
                                    <td class="line-content">
                                        <pre><code class="language-{hljs_lang}">{code}</code></pre>
                                    </td>
                                </tr>
                        """

                    html_template += """
                                </table>
                            </div>
                        </div>
                    """

                # Comments section
                comments = issue.get("comments", [])
                if comments:
                    html_template += """
                        <div class="comments-section">
                            <div class="detail-label" data-i18n-key="comments">Comments</div>
                    """

                    for comment in comments:
                        author = comment.get("login", "Unknown")
                        created_at = comment.get("createdAt", "")
                        html_text = comment.get("htmlText", "")

                        # Format date
                        if created_at:
                            try:
                                dt = datetime.fromisoformat(
                                    created_at.replace("Z", "+00:00")
                                )
                                formatted_date = dt.strftime("%d.%m.%y %H:%M")
                            except:
                                formatted_date = created_at
                        else:
                            formatted_date = "Unknown date"

                        html_template += f"""
                            <div class="comment">
                                <div class="comment-header">
                                    <span>{author}</span>
                                    <span>{formatted_date}</span>
                                </div>
                                <div class="comment-content">{html_text}</div>
                            </div>
                        """

                    html_template += """
                        </div>
                    """

                html_template += """
                    </div>
                </div>
                """

            html_template += """
            </div>
            """

        html_template += """
        </div>
    </div>
    
    <script>
        // Translation dictionary
        const translations = {
            'en': {
                'report-title': 'SonarQube Security Report',
                'report-date-title': 'Generated on:',
                'total-issues': 'Total Issues',
                'bugs': 'Total ',
                'vulnerabilities': 'Vulnerabilities',
                'bugs': 'Bugs',
                'code-smells': 'Code Smells',
                'severity-distribution': 'Severity Distribution',
                'status-distribution': 'Status Distribution',
                'severity': 'Severity',
                'status': 'Status',
                'count': 'Count',
                'percentage': 'Percentage',
                'severity-BLOCKER': 'BLOCKER',
                'severity-CRITICAL': 'CRITICAL',
                'severity-MAJOR': 'MAJOR',
                'severity-MINOR': 'MINOR',
                'severity-INFO': 'INFO',
                'status-CLOSED': 'CLOSED',
                'status-OPEN': 'OPEN',
                'status-RESOLVED': 'RESOLVED',
                'category-header-VULNERABILITY': 'Vulnerabilities',
                'category-header-BUG': 'Bugs',
                'category-header-CODE_SMELL': 'Code smells',
                'issue-description': 'Issue Description',
                'key': 'Key',
                'type': 'Type',
                'author': 'Author',
                'rule': 'Rule',
                'language': 'Language',
                'source-code': 'Source Code',
                'comments': 'Comments'
                // Add more English translations as needed
            },
            'ru': {
                'report-title': 'Отчёт безопасности SonarQube',
                'report-date-title': 'Отчет сгенерирован:',
                'total-issues': 'Всего уязвимостей',
                'vulnerabilities': 'Уязвимости',
                'bugs': 'Баги',
                'code-smells': 'Код с запашком',
                'severity-distribution': 'Распределение по серьёзности',
                'status-distribution': 'Распределение по статусам',
                'severity': 'Серьёзность',
                'status': 'Статус',
                'count': 'Количество',
                'percentage': 'Процент',
                'severity-BLOCKER': 'НАИВЫСШАЯ',
                'severity-CRITICAL': 'КРИТИЧЕСКАЯ',
                'severity-MAJOR': 'ВЫСОКАЯ',
                'severity-MINOR': 'НИЗКАЯ',
                'severity-INFO': 'ИНФО',
                'status-CLOSED': 'Закрыто',
                'status-OPEN': 'Не обработано',
                'status-RESOLVED': 'Решено',
                'category-header-VULNERABILITY': 'Уязвимости',
                'category-header-BUG': 'Баги',
                'category-header-CODE_SMELL': 'Код с запашком',
                'issue-description': 'Описание ошибки',
                'key': 'Ключ-идентификатор',
                'type': 'Тип',
                'author': 'Автор кода',
                'rule': 'Правило',
                'language': 'Язык',
                'source-code': 'Исходный код',
                'comments': 'Комментарии'
                // Add more Russian translations as needed
            }
        };

        let currentLanguage = 'ru';

        function switchLanguage() {
            currentLanguage = currentLanguage === 'en' ? 'ru' : 'en';
            applyTranslations();
        }

        function applyTranslations() {
            // Update all elements with data-i18n-key attribute
            document.querySelectorAll('[data-i18n-key]').forEach(element => {
                const key = element.getAttribute('data-i18n-key');
                if (translations[currentLanguage][key]) {
                    element.textContent = translations[currentLanguage][key];
                }
            });
            
            // Update page title and other special elements
            const titleElement = document.querySelector('title');
            if (titleElement) {
                titleElement.textContent = translations[currentLanguage]['report-title'];
            }
        }

        // Initialize translations when page loads
        document.addEventListener('DOMContentLoaded', function() {
            applyTranslations();
        });
    </script>
    <script>hljs.highlightAll();</script>
</body>
</html>
        """

        # Write HTML to file
        with open(output_filename, "w", encoding="utf-8") as f:
            f.write(html_template)

        print(f"Enhanced HTML report with Highlight.js generated: {output_filename}")

        return {
            "total_issues": total_issues,
            "vulnerabilities": type_counts["VULNERABILITY"],
            "bugs": type_counts["BUG"],
            "code_smells": type_counts["CODE_SMELL"],
            "severity_counts": severity_counts,
            "status_counts": status_counts,
            "html_file": output_filename,
        }

    except Exception as e:
        print(f"Error generating enhanced HTML report: {e}")
        import traceback

        traceback.print_exc()
        return None


def extract_sources_from_response(snippet_data):
    """
    Extract sources from the nested structure where the top key is dynamic
    and clean HTML from code content while preserving formatting
    """
    try:
        if not snippet_data or not isinstance(snippet_data, dict):
            return []

        # Get the first top-level key
        first_key = list(snippet_data.keys())[0]
        component_data = snippet_data[first_key]

        # Extract sources if available
        if (
            isinstance(component_data, dict)
            and "sources" in component_data
            and isinstance(component_data["sources"], list)
        ):

            # Clean HTML from each source's code content while preserving formatting
            sources = []
            for source in component_data["sources"]:
                sources.append(source)

            return sources

        return []

    except (IndexError, KeyError, TypeError) as e:
        print(f"Warning: Could not extract sources from response: {e}")
        return []


def fetch_issue_snippets(all_issues, base_url, cookies, headers):
    """
    Fetch source code snippets for issues that have textRange
    and add them to the corresponding issue objects
    """
    print("Fetching source code snippets for issues...")

    issues_with_snippets = 0
    total_issues = len(all_issues)

    for index, issue in enumerate(all_issues):
        # Check if issue has textRange
        if issue.get("textRange"):
            issue_key = issue.get("key", "")
            if issue_key:
                try:
                    # Build the URL for snippet API
                    snippet_url = (
                        f"{base_url}/api/sources/issue_snippets?issueKey={issue_key}"
                    )

                    # Send GET request
                    response = requests.get(
                        url=snippet_url, headers=headers, cookies=cookies, verify=False
                    )

                    # Check if request was successful
                    if response.status_code == 200:
                        snippet_data = response.json()

                        # Extract sources using the new function
                        sources = extract_sources_from_response(snippet_data)

                        # Add sources to the issue
                        issue["sources"] = sources
                        issues_with_snippets += 1

                        # Print progress
                        if (index + 1) % 50 == 0 or (index + 1) == total_issues:
                            print(
                                f"Processed {index + 1}/{total_issues} issues, fetched snippets for {issues_with_snippets} issues"
                            )

                    else:
                        print(
                            f"Warning: Failed to fetch snippets for issue {issue_key}, status: {response.status_code}"
                        )
                        issue["sources"] = []

                    # Add small delay to avoid overwhelming the server
                    time.sleep(0.1)

                except requests.exceptions.RequestException as e:
                    print(f"Error fetching snippets for issue {issue_key}: {e}")
                    issue["sources"] = []
                except Exception as e:
                    print(f"Unexpected error for issue {issue_key}: {e}")
                    issue["sources"] = []
        else:
            # No textRange, no sources
            issue["sources"] = []

    print(
        f"Successfully fetched snippets for {issues_with_snippets} out of {total_issues} issues"
    )
    return all_issues


def generate_excel_report(all_issues, output_filename="sonarqube_report.xlsx"):
    """Generate Excel report with issues categorized by type"""

    # Categorize issues
    vulnerabilities = []
    bugs = []
    code_smells = []

    for issue in all_issues:
        issue_type = issue.get("type", "").upper()

        if issue_type == "VULNERABILITY":
            vulnerabilities.append(issue)
        elif issue_type == "BUG":
            bugs.append(issue)
        elif issue_type == "CODE_SMELL":
            code_smells.append(issue)

    # Create workbook and sheets
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Create sheets for each issue type
    vuln_sheet = wb.create_sheet("VULNERABILITIES")
    bug_sheet = wb.create_sheet("BUGS")
    smell_sheet = wb.create_sheet("CODE_SMELLS")

    # Define headers
    headers = [
        "Key",
        "External Rule Engine",
        "File",
        "Line",
        "Language",
        "Message",
        "Status",
        "Severity",
        "Resolution",
        "Comment",
        "Comment Author",
        "Code Author",
    ]

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )

    # Helper function to write data to sheet
    def write_issues_to_sheet(sheet, issues):
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # Write issues
        for row, issue in enumerate(issues, 2):
            component = issue.get("component", "")
            file_path = extract_filename(component)

            # Get comment information
            comments = issue.get("comments", [])
            comment_text = ""
            comment_author = ""
            if comments:
                # Get the first comment
                first_comment = comments[0]
                comment_text = first_comment.get("htmlText", "")
                comment_author = first_comment.get("login", "")

            # Get start line information
            line = issue.get("textRange", {})
            start_line = line.get("startLine", "")

            # Prepare row data
            row_data = [
                issue.get("key", ""),
                issue.get("externalRuleEngine", ""),
                file_path,
                start_line,
                get_language_from_extension(file_path),
                issue.get("message", ""),
                issue.get("status", ""),
                issue.get("severity", ""),
                issue.get("resolution", ""),
                comment_text,
                comment_author,
                issue.get("author", ""),  # code author
            ]

            # Write row data
            for col, value in enumerate(row_data, 1):
                sheet.cell(row=row, column=col, value=value)

        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width

        # ADD AUTO-FILTER HERE - This is the key addition
        # Set auto_filter to cover all headers and data (from row 1 to last row with data)
        sheet.auto_filter.ref = f"A1:{sheet.cell(row=1, column=len(headers)).column_letter}{len(issues) + 1}"

    # Write data to each sheet
    write_issues_to_sheet(vuln_sheet, vulnerabilities)
    write_issues_to_sheet(bug_sheet, bugs)
    write_issues_to_sheet(smell_sheet, code_smells)

    # Save workbook
    wb.save(output_filename)
    print(f"Excel report generated: {output_filename}")

    return {
        "vulnerabilities": len(vulnerabilities),
        "bugs": len(bugs),
        "code_smells": len(code_smells),
        "excel_file": output_filename,
    }


def main():
    if len(sys.argv) != 2:
        print("Usage: python script.py <config_file_path>")
        sys.exit(1)

    config_path = sys.argv[1]

    try:
        # Read configuration from JSON file
        with open(config_path, "r") as config_file:
            config = json.load(config_file)

        url = config.get("url")
        project_id = config.get("project_id")
        project_name = config.get("project_name")
        project_version = config.get("project_version")
        branch = config.get("branch")
        jwt_session = config.get("JWT-SESSION")

        if (
            not url
            or not project_id
            or not project_name
            or not project_version
            or not jwt_session
            or not branch
        ):
            print("Error: all fields must be provided in the config file.")
            sys.exit(1)

        # Set up headers and cookies
        headers = {"Accept": "application/json"}
        cookies = {"JWT-SESSION": jwt_session}

        output_file = "response_output.json"
        all_issues = []
        total_issues = 0
        issues_per_page = "500"
        page = 1

        # Send GET request with SSL verification disabled
        while True:
            response = requests.get(
                url=url
                + "/api/issues/search?components="
                + project_id
                + "&branch="
                + branch
                + "&scopes=MAIN"
                + "&impactSeverities=BLOCKER%2CHIGH%2CMEDIUM%2CINFO%2CLOW"
                + "&impactSoftwareQualities=RELIABILITY%2CSECURITY"
                + "&issueStatuses=CONFIRMED%2CFALSE_POSITIVE%2CFIXED%2COPEN"
                + "&ps="
                + issues_per_page
                + "&p="
                + str(page)
                + "&additionalFields=_all",
                headers=headers,
                cookies=cookies,
                verify=False,  # Disable SSL verification
            )
            # Check if request was successful
            response.raise_for_status()

            # Parse JSON response
            response_data = response.json()

            # Extract issues from the current page
            issues_on_page = response_data.get("issues", [])
            print(f"Processing {len(issues_on_page)} issues from page: {page}")

            all_issues.extend(issues_on_page)

            # Update total issues from the first response
            if page == 1:
                total_issues = response_data.get("total", 0)

            if not issues_on_page or page * int(issues_per_page) >= total_issues:
                break  # Exit loop if no more issues are returned

            page += 1

            if page == 21:
                print(
                    f"Reached page limit of 10000 issues from SonarQube API. Try to filter the issues to gen less than 10000 issues."
                )
                sys.exit(1)

        # Fetch source code snippets for issues with textRange
        all_issues = fetch_issue_snippets(all_issues, url, cookies, headers)

        # Save raw JSON response with sources included
        with open(output_file, "w", encoding="utf-8") as outfile:
            json.dump(all_issues, outfile, indent=4, ensure_ascii=False)

        print(f"Success! Response with sources saved to {output_file}")

        # Generate Excel report
        report_stats = generate_excel_report(all_issues, "sonarqube_issues_report.xlsx")
        print(f"Report summary: {report_stats}")

        print("Generating single HTML report with all issues...")
        html_result = generate_single_html_report(
            project_name,
            project_version,
            all_issues,
            "sonarqube_comprehensive_report.html",
        )

        if html_result:
            print(f"HTML report summary: {html_result}")
        else:
            print("Failed to generate HTML report")

    except FileNotFoundError:
        print(f"Error: Config file '{config_path}' not found")
        sys.exit(1)
    except json.JSONDecodeError:
        print("Error: Invalid JSON format in config file")
        sys.exit(1)
    except requests.exceptions.RequestException as e:
        print(f"Error: Request failed - {e}")
        print(f"Check the connection and access to SonarQube - {e}")
        sys.exit(1)
    except Exception as e:
        print(f"Error: {e}")
        import traceback

        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
